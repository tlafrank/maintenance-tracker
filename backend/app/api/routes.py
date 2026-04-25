import logging
import io
from pathlib import Path
from uuid import uuid4
from datetime import datetime, timedelta

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from PIL import Image, UnidentifiedImageError
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy import desc, select
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.core.config import settings
from app.core.security import create_access_token, hash_password, verify_password
from app.db.session import get_db
from app.models.models import Asset, AssetType, MaintenanceEvent, MaintenanceSchedule, MaintenanceTaskTemplate, Meter, MeterReading, User
from app.schemas.schemas import (
    AssetCreate,
    AssetOut,
    AssetTypeCreate,
    AssetTypeOut,
    AssetUpdate,
    DashboardItem,
    DashboardOut,
    MaintenanceEventCreate,
    MaintenanceEventOut,
    MaintenanceActivitySuggestion,
    MaintenanceTaskSuggestion,
    MaintenanceTaskRename,
    MaintenanceTaskCreate,
    MaintenanceTaskDeleteImpact,
    MaintenanceTaskUpdate,
    MeterCreate,
    MeterOut,
    MeterReadingCreate,
    MeterReadingOut,
    ScheduleCreate,
    ScheduleIntervalUpdate,
    ScheduleOut,
    ScheduleUpdate,
    Token,
    UserCreate,
    UserProfileUpdate,
    UserOut,
)
from app.services.due_logic import evaluate_schedule_status, latest_meter_map

router = APIRouter()
auth_logger = logging.getLogger('app.auth')
UPLOAD_DIRECTORY = Path(settings.uploads_dir) / 'assets'
ALLOWED_IMAGE_TYPES = {'image/jpeg', 'image/png', 'image/webp', 'image/gif'}
MAX_IMAGE_BYTES = 5 * 1024 * 1024
MAX_THUMBNAIL_DIMENSIONS = (1024, 1024)


@router.post('/auth/register', response_model=UserOut)
def register(payload: UserCreate, request: Request, db: Session = Depends(get_db)):
    existing = db.scalar(select(User).where(User.email == payload.email))
    if existing:
        auth_logger.warning(
            'register_failed_existing_email email=%s ip=%s',
            payload.email,
            request.client.host if request.client else 'unknown',
        )
        raise HTTPException(status_code=400, detail='Email already registered')
    user = User(
        email=payload.email,
        display_name=(payload.display_name.strip() if payload.display_name else payload.email.split('@')[0]),
        password_hash=hash_password(payload.password),
        preferred_distance_unit='km',
        upcoming_task_window_days=14,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    auth_logger.info(
        'register_success user_id=%s email=%s ip=%s',
        user.id,
        user.email,
        request.client.host if request.client else 'unknown',
    )
    return user


@router.post('/auth/login', response_model=Token)
def login(request: Request, form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.scalar(select(User).where(User.email == form_data.username))
    if not user or not verify_password(form_data.password, user.password_hash):
        auth_logger.warning(
            'login_failed email=%s ip=%s user_agent=%s',
            form_data.username,
            request.client.host if request.client else 'unknown',
            request.headers.get('user-agent', 'unknown'),
        )
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail='Invalid credentials')
    auth_logger.info(
        'login_success user_id=%s email=%s ip=%s user_agent=%s',
        user.id,
        user.email,
        request.client.host if request.client else 'unknown',
        request.headers.get('user-agent', 'unknown'),
    )
    return Token(access_token=create_access_token(str(user.id)))


@router.get('/auth/me', response_model=UserOut)
def me(current_user: User = Depends(get_current_user)):
    return current_user


@router.put('/auth/profile', response_model=UserOut)
def update_profile(payload: UserProfileUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    current_user.preferred_distance_unit = payload.preferred_distance_unit.strip() or 'km'
    current_user.upcoming_task_window_days = max(1, int(payload.upcoming_task_window_days or 14))

    if payload.new_password:
        if not payload.current_password or not verify_password(payload.current_password, current_user.password_hash):
            raise HTTPException(status_code=400, detail='Current password is incorrect')
        current_user.password_hash = hash_password(payload.new_password)

    db.commit()
    db.refresh(current_user)
    return current_user


@router.get('/auth/profile/export')
def export_profile_workbook(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    workbook = Workbook()
    assets_sheet = workbook.active
    assets_sheet.title = 'Assets'
    assets_sheet.append([
        'asset_external_id',
        'name',
        'asset_type',
        'manufacturer',
        'model',
        'year',
        'registration_or_serial',
        'notes',
        'service_trigger',
        'meter_current_value',
        'archived_at',
    ])

    events_sheet = workbook.create_sheet('Maintenance Activities')
    events_sheet.append([
        'asset_external_id',
        'asset_name',
        'performed_at',
        'event_type',
        'notes',
        'completion_meter_value',
        'schedule_title',
    ])

    schedules_sheet = workbook.create_sheet('Scheduled Tasks')
    schedules_sheet.append([
        'asset_external_id',
        'asset_name',
        'title',
        'description',
        'interval_days',
        'interval_distance',
        'interval_hours',
        'interval_cycles',
        'due_soon_threshold_days',
        'due_soon_threshold_distance',
        'due_soon_threshold_hours',
        'due_soon_threshold_cycles',
        'active',
    ])

    assets = db.scalars(
        select(Asset).where(Asset.owner_user_id == current_user.id).order_by(Asset.id)
    ).all()

    meters_by_asset = {
        meter.asset_id: meter
        for meter in db.scalars(select(Meter).join(Asset, Meter.asset_id == Asset.id).where(Asset.owner_user_id == current_user.id)).all()
    }
    asset_by_id = {asset.id: asset for asset in assets}

    for asset in assets:
        meter = meters_by_asset.get(asset.id)
        assets_sheet.append([
            asset.id,
            asset.name,
            asset.asset_type,
            asset.manufacturer,
            asset.model,
            asset.year,
            asset.registration_or_serial,
            asset.notes,
            asset.service_trigger,
            float(meter.current_value) if meter and meter.current_value is not None else None,
            asset.archived_at.isoformat() if asset.archived_at else None,
        ])

    schedules = db.scalars(
        select(MaintenanceSchedule).join(Asset, MaintenanceSchedule.asset_id == Asset.id).where(Asset.owner_user_id == current_user.id).order_by(MaintenanceSchedule.asset_id, MaintenanceSchedule.id)
    ).all()
    for schedule in schedules:
        asset = asset_by_id.get(schedule.asset_id)
        if not asset:
            continue
        schedules_sheet.append([
            asset.id,
            asset.name,
            schedule.title,
            schedule.description,
            schedule.interval_days,
            schedule.interval_distance,
            schedule.interval_hours,
            schedule.interval_cycles,
            schedule.due_soon_threshold_days,
            schedule.due_soon_threshold_distance,
            schedule.due_soon_threshold_hours,
            schedule.due_soon_threshold_cycles,
            bool(schedule.active),
        ])

    events = db.scalars(
        select(MaintenanceEvent).join(Asset, MaintenanceEvent.asset_id == Asset.id).where(Asset.owner_user_id == current_user.id).order_by(MaintenanceEvent.asset_id, MaintenanceEvent.performed_at)
    ).all()
    for event in events:
        asset = asset_by_id.get(event.asset_id)
        if not asset:
            continue
        events_sheet.append([
            asset.id,
            asset.name,
            event.performed_at.isoformat() if event.performed_at else None,
            event.event_type,
            event.notes,
            event.completion_meter_value,
            event.schedule.title if event.schedule else None,
        ])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    timestamp = datetime.utcnow().strftime('%Y%m%d-%H%M%S')
    headers = {
        'Content-Disposition': f'attachment; filename="maintenance-profile-export-{timestamp}.xlsx"',
    }
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers,
    )


@router.post('/auth/profile/import')
async def import_profile_workbook(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not file.filename or not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(status_code=400, detail='Upload an .xlsx workbook')
    workbook_bytes = await file.read()
    try:
        workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail='Invalid workbook file') from exc

    required_sheets = {'Assets', 'Maintenance Activities', 'Scheduled Tasks'}
    missing_sheets = [name for name in required_sheets if name not in workbook.sheetnames]
    if missing_sheets:
        raise HTTPException(status_code=400, detail=f'Missing required sheet(s): {", ".join(missing_sheets)}')

    assets_sheet = workbook['Assets']
    events_sheet = workbook['Maintenance Activities']
    schedules_sheet = workbook['Scheduled Tasks']

    def rows_as_dicts(sheet):
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
        output_rows = []
        for row in rows[1:]:
            if not row or all(cell in (None, '') for cell in row):
                continue
            output_rows.append({headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))})
        return output_rows

    asset_rows = rows_as_dicts(assets_sheet)
    schedule_rows = rows_as_dicts(schedules_sheet)
    event_rows = rows_as_dicts(events_sheet)

    asset_refs: dict[str, Asset] = {}
    created_assets = 0
    created_schedules = 0
    created_events = 0
    created_templates = 0
    imported_types: set[str] = set()

    existing_types = {item.name for item in db.scalars(select(AssetType)).all()}
    existing_templates = {
        (template.asset_type or '', template.task_name.strip().lower())
        for template in db.scalars(select(MaintenanceTaskTemplate).where(MaintenanceTaskTemplate.owner_user_id == current_user.id)).all()
    }

    for row in asset_rows:
        asset_name = _xlsx_string(row.get('name'))
        asset_type_name = _xlsx_string(row.get('asset_type'))
        if not asset_name or not asset_type_name:
            continue
        if asset_type_name not in existing_types:
            db.add(AssetType(name=asset_type_name))
            existing_types.add(asset_type_name)
            imported_types.add(asset_type_name)
        asset = Asset(
            owner_user_id=current_user.id,
            name=asset_name,
            asset_type=asset_type_name,
            manufacturer=_xlsx_string(row.get('manufacturer')),
            model=_xlsx_string(row.get('model')),
            year=_xlsx_int(row.get('year')),
            registration_or_serial=_xlsx_string(row.get('registration_or_serial')),
            notes=_xlsx_string(row.get('notes')),
            service_trigger=_xlsx_string(row.get('service_trigger')) or 'distance',
            archived_at=_xlsx_datetime(row.get('archived_at')),
        )
        db.add(asset)
        db.flush()
        created_assets += 1

        external_id = _xlsx_string(str(row.get('asset_external_id')) if row.get('asset_external_id') is not None else None)
        asset_refs[external_id or f'name:{asset_name}'] = asset
        asset_refs[f'name:{asset_name}'] = asset

        meter_value = _xlsx_float(row.get('meter_current_value'))
        if meter_value is not None:
            db.add(
                Meter(
                    asset_id=asset.id,
                    service_trigger=asset.service_trigger,
                    unit=_unit_for_trigger(asset.service_trigger, current_user.preferred_distance_unit),
                    current_value=meter_value,
                )
            )

    schedule_map: dict[tuple[int, str], MaintenanceSchedule] = {}

    for row in schedule_rows:
        asset_external_id = _xlsx_string(str(row.get('asset_external_id')) if row.get('asset_external_id') is not None else None)
        asset_name = _xlsx_string(row.get('asset_name'))
        title = _xlsx_string(row.get('title'))
        if not title:
            continue
        asset = asset_refs.get(asset_external_id or '')
        if not asset and asset_name:
            asset = asset_refs.get(f'name:{asset_name}')
        if not asset:
            continue
        schedule = MaintenanceSchedule(
            asset_id=asset.id,
            title=title,
            description=_xlsx_string(row.get('description')),
            interval_days=_xlsx_int(row.get('interval_days')),
            interval_distance=_xlsx_float(row.get('interval_distance')),
            interval_hours=_xlsx_float(row.get('interval_hours')),
            interval_cycles=_xlsx_float(row.get('interval_cycles')),
            due_soon_threshold_days=_xlsx_int(row.get('due_soon_threshold_days')),
            due_soon_threshold_distance=_xlsx_float(row.get('due_soon_threshold_distance')),
            due_soon_threshold_hours=_xlsx_float(row.get('due_soon_threshold_hours')),
            due_soon_threshold_cycles=_xlsx_float(row.get('due_soon_threshold_cycles')),
            active=_xlsx_bool(row.get('active'), default=True),
        )
        db.add(schedule)
        created_schedules += 1
        db.flush()
        schedule_map[(asset.id, title.strip().lower())] = schedule

        template_key = (asset.asset_type or '', title.strip().lower())
        if template_key not in existing_templates:
            db.add(MaintenanceTaskTemplate(owner_user_id=current_user.id, asset_type=asset.asset_type, task_name=title))
            existing_templates.add(template_key)
            created_templates += 1

    for row in event_rows:
        asset_external_id = _xlsx_string(str(row.get('asset_external_id')) if row.get('asset_external_id') is not None else None)
        asset_name = _xlsx_string(row.get('asset_name'))
        event_type = _xlsx_string(row.get('event_type'))
        if not event_type:
            continue
        asset = asset_refs.get(asset_external_id or '')
        if not asset and asset_name:
            asset = asset_refs.get(f'name:{asset_name}')
        if not asset:
            continue
        schedule_title = _xlsx_string(row.get('schedule_title'))
        schedule = schedule_map.get((asset.id, schedule_title.strip().lower())) if schedule_title else None
        event = MaintenanceEvent(
            asset_id=asset.id,
            schedule_id=schedule.id if schedule else None,
            performed_by_user_id=current_user.id,
            performed_at=_xlsx_datetime(row.get('performed_at')) or datetime.utcnow(),
            event_type=event_type,
            notes=_xlsx_string(row.get('notes')),
            completion_meter_value=_xlsx_float(row.get('completion_meter_value')),
        )
        db.add(event)
        created_events += 1
        meter_value = _xlsx_float(row.get('completion_meter_value'))
        if meter_value is not None:
            meter = db.scalar(select(Meter).where(Meter.asset_id == asset.id).order_by(desc(Meter.id)))
            if not meter:
                meter = Meter(
                    asset_id=asset.id,
                    service_trigger=asset.service_trigger,
                    unit=_unit_for_trigger(asset.service_trigger, current_user.preferred_distance_unit),
                    current_value=meter_value,
                )
                db.add(meter)
                db.flush()
            else:
                meter.current_value = meter_value
            db.add(
                MeterReading(
                    meter_id=meter.id,
                    asset_id=asset.id,
                    reading_value=meter_value,
                    reading_timestamp=event.performed_at,
                    source='import',
                    notes='Imported from profile workbook',
                )
            )

        for task_name in [task.strip() for task in event_type.split(',') if task.strip()]:
            template_key = (asset.asset_type or '', task_name.lower())
            if template_key in existing_templates:
                continue
            db.add(MaintenanceTaskTemplate(owner_user_id=current_user.id, asset_type=asset.asset_type, task_name=task_name))
            existing_templates.add(template_key)
            created_templates += 1

    db.commit()
    return {
        'imported_asset_types': len(imported_types),
        'imported_assets': created_assets,
        'imported_schedules': created_schedules,
        'imported_events': created_events,
        'imported_task_templates': created_templates,
    }




@router.get('/asset-types', response_model=list[AssetTypeOut])
def list_asset_types(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return db.scalars(select(AssetType).order_by(AssetType.name)).all()


@router.post('/asset-types', response_model=AssetTypeOut)
def create_asset_type(payload: AssetTypeCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    normalized_name = payload.name.strip()
    if not normalized_name:
        raise HTTPException(status_code=400, detail='Asset type name is required')

    existing = db.scalar(select(AssetType).where(AssetType.name == normalized_name))
    if existing:
        raise HTTPException(status_code=400, detail='Asset type already exists')

    asset_type = AssetType(name=normalized_name)
    db.add(asset_type)
    db.commit()
    db.refresh(asset_type)
    return asset_type


@router.get('/assets', response_model=list[AssetOut])
def list_assets(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return db.scalars(select(Asset).where(Asset.owner_user_id == current_user.id, Asset.archived_at.is_(None))).all()


@router.post('/assets', response_model=AssetOut)
def create_asset(payload: AssetCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    asset_type = db.scalar(select(AssetType).where(AssetType.name == payload.asset_type))
    if not asset_type:
        raise HTTPException(status_code=400, detail='Select a valid asset type')

    asset = Asset(owner_user_id=current_user.id, **payload.model_dump())
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return asset


def _owned_asset(asset_id: int, user_id: int, db: Session) -> Asset:
    asset = db.get(Asset, asset_id)
    if not asset or asset.owner_user_id != user_id or asset.archived_at is not None:
        raise HTTPException(status_code=404, detail='Asset not found')
    return asset


def _delete_existing_thumbnail(asset: Asset):
    if not asset.thumbnail_path:
        return
    existing_path = Path(settings.uploads_dir) / asset.thumbnail_path.replace('/uploads/', '', 1).lstrip('/')
    if existing_path.exists():
        existing_path.unlink()


def _validate_service_interval(interval_days: int | None, usage_interval: float | None):
    if interval_days is None and usage_interval is None:
        raise HTTPException(status_code=400, detail='Service interval requires time, service trigger usage, or both.')


def _usage_interval_for_trigger(asset: Asset, payload: ScheduleCreate | ScheduleUpdate | ScheduleIntervalUpdate) -> float | None:
    if asset.service_trigger == 'distance':
        return payload.interval_distance
    if asset.service_trigger == 'hours':
        return payload.interval_hours
    if asset.service_trigger == 'cycles':
        return payload.interval_cycles
    return None


def _normalize_usage_intervals(asset: Asset, payload: dict) -> dict:
    payload['interval_distance'] = payload.get('interval_distance') if asset.service_trigger == 'distance' else None
    payload['interval_hours'] = payload.get('interval_hours') if asset.service_trigger == 'hours' else None
    payload['interval_cycles'] = payload.get('interval_cycles') if asset.service_trigger == 'cycles' else None
    return payload


def _unit_for_trigger(service_trigger: str, preferred_distance_unit: str = 'km') -> str:
    if service_trigger == 'distance':
        return preferred_distance_unit or 'km'
    if service_trigger == 'hours':
        return 'h'
    if service_trigger == 'cycles':
        return 'cycles'
    return ''


def _xlsx_string(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip()
    return normalized or None


def _xlsx_datetime(value) -> datetime | None:
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.strip())
        except ValueError:
            return None
    return None


def _xlsx_float(value) -> float | None:
    if value in (None, ''):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _xlsx_int(value) -> int | None:
    if value in (None, ''):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _xlsx_bool(value, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {'true', '1', 'yes', 'y'}:
            return True
        if normalized in {'false', '0', 'no', 'n'}:
            return False
    return default


@router.get('/assets/{asset_id}', response_model=AssetOut)
def get_asset(asset_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return _owned_asset(asset_id, current_user.id, db)


@router.put('/assets/{asset_id}', response_model=AssetOut)
def update_asset(asset_id: int, payload: AssetUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    asset = _owned_asset(asset_id, current_user.id, db)
    asset_type = db.scalar(select(AssetType).where(AssetType.name == payload.asset_type))
    if not asset_type:
        raise HTTPException(status_code=400, detail='Select a valid asset type')
    for key, value in payload.model_dump().items():
        setattr(asset, key, value)
    db.commit()
    db.refresh(asset)
    return asset


@router.post('/assets/{asset_id}/thumbnail', response_model=AssetOut)
async def upload_asset_thumbnail(
    asset_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    asset = _owned_asset(asset_id, current_user.id, db)
    if file.content_type not in ALLOWED_IMAGE_TYPES:
        raise HTTPException(status_code=400, detail='Upload a JPG, PNG, WEBP, or GIF image.')

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail='Uploaded image is empty.')
    if len(content) > MAX_IMAGE_BYTES:
        raise HTTPException(status_code=400, detail='Image must be 5MB or less.')

    try:
        source_image = Image.open(io.BytesIO(content))
    except UnidentifiedImageError as exc:
        raise HTTPException(status_code=400, detail='Uploaded file is not a valid image.') from exc

    has_alpha_channel = source_image.mode in {'RGBA', 'LA'} or 'transparency' in source_image.info
    optimized_image = source_image.convert('RGBA' if has_alpha_channel else 'RGB')
    optimized_image.thumbnail(MAX_THUMBNAIL_DIMENSIONS)
    optimized_buffer = io.BytesIO()
    save_kwargs = {'format': 'WEBP', 'method': 6}
    if has_alpha_channel:
        save_kwargs['lossless'] = True
    else:
        save_kwargs['quality'] = 82
    optimized_image.save(optimized_buffer, **save_kwargs)
    optimized_content = optimized_buffer.getvalue()

    extension = '.webp'

    user_upload_directory = UPLOAD_DIRECTORY / str(current_user.id)
    user_upload_directory.mkdir(parents=True, exist_ok=True)
    filename = f'{asset.id}-{uuid4().hex}{extension}'
    saved_path = user_upload_directory / filename
    saved_path.write_bytes(optimized_content)

    _delete_existing_thumbnail(asset)
    asset.thumbnail_path = f'/uploads/assets/{current_user.id}/{filename}'
    db.commit()
    db.refresh(asset)
    return asset


@router.delete('/assets/{asset_id}')
def archive_asset(asset_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    asset = _owned_asset(asset_id, current_user.id, db)
    asset.archived_at = datetime.utcnow()
    db.commit()
    return {'ok': True}


@router.post('/assets/{asset_id}/meters', response_model=MeterOut)
def create_meter(asset_id: int, payload: MeterCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    asset = _owned_asset(asset_id, current_user.id, db)
    if payload.service_trigger != asset.service_trigger:
        raise HTTPException(status_code=400, detail='Reading source must use the asset Service Trigger.')
    existing_meter = db.scalar(
        select(Meter)
        .where(Meter.asset_id == asset_id)
        .order_by(desc(Meter.id))
    )
    if existing_meter:
        existing_meter.service_trigger = asset.service_trigger
        existing_meter.unit = payload.unit
        existing_meter.current_value = payload.current_value
        db.commit()
        db.refresh(existing_meter)
        return existing_meter

    meter = Meter(
        asset_id=asset_id,
        service_trigger=asset.service_trigger,
        unit=payload.unit,
        current_value=payload.current_value,
    )
    db.add(meter)
    db.commit()
    db.refresh(meter)
    return meter


@router.get('/assets/{asset_id}/meters', response_model=list[MeterOut])
def list_meters(asset_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _owned_asset(asset_id, current_user.id, db)
    meter = db.scalar(select(Meter).where(Meter.asset_id == asset_id).order_by(desc(Meter.id)))
    return [meter] if meter else []


@router.post('/assets/{asset_id}/readings', response_model=MeterReadingOut)
def create_reading(asset_id: int, payload: MeterReadingCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    asset = _owned_asset(asset_id, current_user.id, db)
    meter = db.get(Meter, payload.meter_id) if payload.meter_id else db.scalar(
        select(Meter).where(Meter.asset_id == asset_id).order_by(desc(Meter.id))
    )
    if not meter:
        meter = Meter(
            asset_id=asset_id,
            service_trigger=asset.service_trigger,
            unit=_unit_for_trigger(asset.service_trigger),
            current_value=payload.reading_value,
        )
        db.add(meter)
        db.flush()
    elif meter.asset_id != asset_id:
        raise HTTPException(status_code=404, detail='Meter not found')
    reading = MeterReading(
        meter_id=meter.id,
        asset_id=asset_id,
        reading_value=payload.reading_value,
        reading_timestamp=payload.reading_timestamp or datetime.utcnow(),
        notes=payload.notes,
    )
    meter.current_value = payload.reading_value
    db.add(reading)
    db.commit()
    db.refresh(reading)
    return reading


@router.get('/assets/{asset_id}/readings', response_model=list[MeterReadingOut])
def list_readings(asset_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _owned_asset(asset_id, current_user.id, db)
    return db.scalars(select(MeterReading).where(MeterReading.asset_id == asset_id).order_by(desc(MeterReading.reading_timestamp))).all()


@router.get('/assets/{asset_id}/schedules', response_model=list[ScheduleOut])
def list_schedules(asset_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _owned_asset(asset_id, current_user.id, db)
    return db.scalars(
        select(MaintenanceSchedule).where(
            MaintenanceSchedule.asset_id == asset_id,
            MaintenanceSchedule.active.is_(True),
        )
    ).all()


@router.post('/assets/{asset_id}/schedules', response_model=ScheduleOut)
def create_schedule(asset_id: int, payload: ScheduleCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    asset = _owned_asset(asset_id, current_user.id, db)
    usage_interval = _usage_interval_for_trigger(asset, payload)
    _validate_service_interval(payload.interval_days, usage_interval)
    schedule_payload = _normalize_usage_intervals(asset, payload.model_dump())
    schedule = MaintenanceSchedule(asset_id=asset_id, **schedule_payload)
    db.add(schedule)
    existing_template = db.scalar(
        select(MaintenanceTaskTemplate).where(
            MaintenanceTaskTemplate.owner_user_id == current_user.id,
            MaintenanceTaskTemplate.asset_type == asset.asset_type,
            MaintenanceTaskTemplate.task_name == payload.title.strip(),
        )
    )
    if not existing_template and payload.title.strip():
        db.add(MaintenanceTaskTemplate(
            owner_user_id=current_user.id,
            asset_type=asset.asset_type,
            task_name=payload.title.strip(),
        ))
    db.commit()
    db.refresh(schedule)
    return schedule


@router.put('/schedules/{schedule_id}', response_model=ScheduleOut)
def update_schedule(schedule_id: int, payload: ScheduleUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    schedule = db.get(MaintenanceSchedule, schedule_id)
    if not schedule:
        raise HTTPException(status_code=404, detail='Schedule not found')
    asset = _owned_asset(schedule.asset_id, current_user.id, db)
    usage_interval = _usage_interval_for_trigger(asset, payload)
    _validate_service_interval(payload.interval_days, usage_interval)
    normalized_payload = _normalize_usage_intervals(asset, payload.model_dump())
    for key, value in normalized_payload.items():
        setattr(schedule, key, value)
    db.commit()
    db.refresh(schedule)
    return schedule


@router.patch('/schedules/{schedule_id}/intervals', response_model=ScheduleOut)
def update_schedule_intervals(schedule_id: int, payload: ScheduleIntervalUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    schedule = db.get(MaintenanceSchedule, schedule_id)
    if not schedule:
        raise HTTPException(status_code=404, detail='Schedule not found')
    asset = _owned_asset(schedule.asset_id, current_user.id, db)
    usage_interval = _usage_interval_for_trigger(asset, payload)
    _validate_service_interval(payload.interval_days, usage_interval)
    normalized_payload = _normalize_usage_intervals(asset, payload.model_dump())

    schedule.interval_days = normalized_payload['interval_days']
    schedule.interval_distance = normalized_payload['interval_distance']
    schedule.interval_hours = normalized_payload['interval_hours']
    schedule.interval_cycles = normalized_payload['interval_cycles']
    db.commit()
    db.refresh(schedule)
    return schedule


@router.delete('/schedules/{schedule_id}')
def delete_schedule(schedule_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    schedule = db.get(MaintenanceSchedule, schedule_id)
    if not schedule:
        raise HTTPException(status_code=404, detail='Schedule not found')
    _owned_asset(schedule.asset_id, current_user.id, db)
    schedule.active = False
    db.commit()
    return {'ok': True}


@router.get('/assets/{asset_id}/maintenance-events', response_model=list[MaintenanceEventOut])
def list_events(asset_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _owned_asset(asset_id, current_user.id, db)
    return db.scalars(select(MaintenanceEvent).where(MaintenanceEvent.asset_id == asset_id).order_by(desc(MaintenanceEvent.performed_at))).all()


@router.post('/assets/{asset_id}/maintenance-events', response_model=MaintenanceEventOut)
def create_event(asset_id: int, payload: MaintenanceEventCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    asset = _owned_asset(asset_id, current_user.id, db)
    if payload.schedule_id:
        schedule = db.get(MaintenanceSchedule, payload.schedule_id)
        if not schedule or schedule.asset_id != asset_id:
            raise HTTPException(status_code=404, detail='Schedule not found')
    event = MaintenanceEvent(
        asset_id=asset_id,
        schedule_id=payload.schedule_id,
        performed_by_user_id=current_user.id,
        performed_at=payload.performed_at or datetime.utcnow(),
        event_type=payload.event_type,
        notes=payload.notes,
        completion_meter_value=payload.completion_meter_value,
    )
    db.add(event)
    event_tasks = [task.strip() for task in payload.event_type.split(',') if task.strip()]
    existing_templates = db.scalars(
        select(MaintenanceTaskTemplate).where(
            MaintenanceTaskTemplate.owner_user_id == current_user.id,
            MaintenanceTaskTemplate.asset_type == asset.asset_type,
        )
    ).all()
    existing_task_names = {template.task_name.strip().lower() for template in existing_templates}
    for task_name in event_tasks:
        if task_name.lower() in existing_task_names:
            continue
        db.add(MaintenanceTaskTemplate(owner_user_id=current_user.id, asset_type=asset.asset_type, task_name=task_name))
        existing_task_names.add(task_name.lower())
    if payload.completion_meter_value is not None:
        meter = db.scalar(select(Meter).where(Meter.asset_id == asset_id).order_by(desc(Meter.id)))
        if not meter:
            meter = Meter(
                asset_id=asset_id,
                service_trigger=asset.service_trigger,
                unit=_unit_for_trigger(asset.service_trigger, current_user.preferred_distance_unit),
                current_value=payload.completion_meter_value,
            )
            db.add(meter)
            db.flush()
        reading = MeterReading(
            meter_id=meter.id,
            asset_id=asset_id,
            reading_value=payload.completion_meter_value,
            notes='Auto-captured from maintenance activity',
        )
        meter.current_value = payload.completion_meter_value
        db.add(reading)
    db.commit()
    db.refresh(event)
    return event


@router.put('/maintenance-events/{event_id}', response_model=MaintenanceEventOut)
def update_event(event_id: int, payload: MaintenanceEventCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    event = db.get(MaintenanceEvent, event_id)
    if not event:
        raise HTTPException(status_code=404, detail='Maintenance event not found')
    asset = _owned_asset(event.asset_id, current_user.id, db)
    event.event_type = payload.event_type
    event.notes = payload.notes
    event.completion_meter_value = payload.completion_meter_value
    if payload.performed_at:
        event.performed_at = payload.performed_at

    if payload.completion_meter_value is not None:
        meter = db.scalar(select(Meter).where(Meter.asset_id == asset.id).order_by(desc(Meter.id)))
        if not meter:
            meter = Meter(
                asset_id=asset.id,
                service_trigger=asset.service_trigger,
                unit=_unit_for_trigger(asset.service_trigger, current_user.preferred_distance_unit),
                current_value=payload.completion_meter_value,
            )
            db.add(meter)
            db.flush()
        else:
            meter.current_value = payload.completion_meter_value
        db.add(MeterReading(
            meter_id=meter.id,
            asset_id=asset.id,
            reading_value=payload.completion_meter_value,
            notes='Auto-captured from maintenance history update',
        ))
    event_tasks = [task.strip() for task in payload.event_type.split(',') if task.strip()]
    existing_templates = db.scalars(
        select(MaintenanceTaskTemplate).where(
            MaintenanceTaskTemplate.owner_user_id == current_user.id,
            MaintenanceTaskTemplate.asset_type == asset.asset_type,
        )
    ).all()
    existing_task_names = {template.task_name.strip().lower() for template in existing_templates}
    for task_name in event_tasks:
        if task_name.lower() in existing_task_names:
            continue
        db.add(MaintenanceTaskTemplate(owner_user_id=current_user.id, asset_type=asset.asset_type, task_name=task_name))
        existing_task_names.add(task_name.lower())
    db.commit()
    db.refresh(event)
    return event


@router.delete('/maintenance-events/{event_id}')
def delete_event(event_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    event = db.get(MaintenanceEvent, event_id)
    if not event:
        raise HTTPException(status_code=404, detail='Maintenance event not found')
    _owned_asset(event.asset_id, current_user.id, db)
    db.delete(event)
    db.commit()
    return {'ok': True}


@router.get('/maintenance-activities', response_model=list[MaintenanceActivitySuggestion])
def list_maintenance_activities(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    events = db.scalars(
        select(MaintenanceEvent)
        .where(MaintenanceEvent.performed_by_user_id == current_user.id)
        .order_by(desc(MaintenanceEvent.performed_at))
    ).all()

    suggestions: list[MaintenanceActivitySuggestion] = []
    seen: set[str] = set()
    for event in events:
        key = event.event_type.strip().lower()
        if not key or key in seen:
            continue
        seen.add(key)
        suggestions.append(
            MaintenanceActivitySuggestion(
                activity_name=event.event_type,
                last_performed_at=event.performed_at,
                last_notes=event.notes,
                last_completion_meter_value=event.completion_meter_value,
            )
        )
    return suggestions


@router.get('/maintenance-tasks', response_model=list[MaintenanceTaskSuggestion])
def list_maintenance_tasks(asset_type: str | None = None, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    normalized_asset_type = asset_type.strip() if asset_type else None
    templates = db.scalars(
        select(MaintenanceTaskTemplate)
        .where(MaintenanceTaskTemplate.owner_user_id == current_user.id)
        .where(MaintenanceTaskTemplate.asset_type == normalized_asset_type if normalized_asset_type else True)
        .order_by(MaintenanceTaskTemplate.task_name)
    ).all()
    tasks: list[MaintenanceTaskSuggestion] = [
        MaintenanceTaskSuggestion(id=template.id, asset_type=template.asset_type, task_name=template.task_name)
        for template in templates
    ]
    return tasks


@router.post('/maintenance-tasks', response_model=MaintenanceTaskSuggestion)
def create_maintenance_task(payload: MaintenanceTaskCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    task_name = payload.task_name.strip()
    normalized_asset_type = payload.asset_type.strip() if payload.asset_type else None
    if not task_name:
        raise HTTPException(status_code=400, detail='Task name is required')
    existing = db.scalar(
        select(MaintenanceTaskTemplate).where(
            MaintenanceTaskTemplate.owner_user_id == current_user.id,
            MaintenanceTaskTemplate.task_name == task_name,
            MaintenanceTaskTemplate.asset_type == normalized_asset_type,
        )
    )
    if existing:
        raise HTTPException(status_code=400, detail='Task already exists')
    template = MaintenanceTaskTemplate(owner_user_id=current_user.id, asset_type=normalized_asset_type, task_name=task_name)
    db.add(template)
    db.commit()
    db.refresh(template)
    return MaintenanceTaskSuggestion(id=template.id, asset_type=template.asset_type, task_name=template.task_name)


@router.put('/maintenance-tasks/{task_id}', response_model=MaintenanceTaskSuggestion)
def update_maintenance_task(task_id: int, payload: MaintenanceTaskUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    template = db.get(MaintenanceTaskTemplate, task_id)
    if not template or template.owner_user_id != current_user.id:
        raise HTTPException(status_code=404, detail='Maintenance task not found')
    previous_task_name = template.task_name.strip()
    previous_asset_type = template.asset_type
    task_name = payload.task_name.strip()
    normalized_asset_type = payload.asset_type.strip() if payload.asset_type else None
    if not task_name:
        raise HTTPException(status_code=400, detail='Task name is required')
    existing = db.scalar(
        select(MaintenanceTaskTemplate).where(
            MaintenanceTaskTemplate.owner_user_id == current_user.id,
            MaintenanceTaskTemplate.task_name == task_name,
            MaintenanceTaskTemplate.asset_type == normalized_asset_type,
            MaintenanceTaskTemplate.id != template.id,
        )
    )
    if existing:
        raise HTTPException(status_code=400, detail='Task already exists')

    matching_assets = db.scalars(
        select(Asset).where(
            Asset.owner_user_id == current_user.id,
            Asset.asset_type == previous_asset_type,
        )
    ).all()
    matching_asset_ids = {asset.id for asset in matching_assets}
    normalized_previous_name = previous_task_name.lower()
    normalized_new_name = task_name.lower()

    if matching_asset_ids and normalized_previous_name != normalized_new_name:
        schedules = db.scalars(select(MaintenanceSchedule).where(MaintenanceSchedule.asset_id.in_(matching_asset_ids))).all()
        for schedule in schedules:
            if schedule.title.strip().lower() == normalized_previous_name:
                schedule.title = task_name

        events = db.scalars(select(MaintenanceEvent).where(MaintenanceEvent.asset_id.in_(matching_asset_ids))).all()
        for event in events:
            tasks = [part.strip() for part in event.event_type.split(',') if part.strip()]
            replaced_tasks = [task_name if task.lower() == normalized_previous_name else task for task in tasks]
            if tasks != replaced_tasks:
                event.event_type = ', '.join(replaced_tasks)

    template.task_name = task_name
    template.asset_type = normalized_asset_type
    db.commit()
    db.refresh(template)
    return MaintenanceTaskSuggestion(id=template.id, asset_type=template.asset_type, task_name=template.task_name)


def _maintenance_task_impact(template: MaintenanceTaskTemplate, current_user: User, db: Session) -> MaintenanceTaskDeleteImpact:
    matching_assets = db.scalars(
        select(Asset).where(
            Asset.owner_user_id == current_user.id,
            Asset.asset_type == template.asset_type,
        )
    ).all()
    matching_asset_ids = {asset.id for asset in matching_assets}

    schedules = db.scalars(
        select(MaintenanceSchedule).where(MaintenanceSchedule.asset_id.in_(matching_asset_ids))
    ).all() if matching_asset_ids else []
    affected_schedules = sum(1 for schedule in schedules if schedule.title.strip().lower() == template.task_name.strip().lower() and schedule.active)

    events = db.scalars(
        select(MaintenanceEvent).where(MaintenanceEvent.asset_id.in_(matching_asset_ids))
    ).all() if matching_asset_ids else []
    affected_history_records = 0
    deleted_history_records = 0
    for event in events:
        tasks = [part.strip() for part in event.event_type.split(',') if part.strip()]
        kept_tasks = [task for task in tasks if task.lower() != template.task_name.strip().lower()]
        if len(kept_tasks) != len(tasks):
            if kept_tasks:
                affected_history_records += 1
            else:
                deleted_history_records += 1

    return MaintenanceTaskDeleteImpact(
        task_name=template.task_name,
        asset_type=template.asset_type,
        affected_schedules=affected_schedules,
        affected_history_records=affected_history_records,
        deleted_history_records=deleted_history_records,
    )


@router.get('/maintenance-tasks/{task_id}/impact', response_model=MaintenanceTaskDeleteImpact)
def maintenance_task_impact(task_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    template = db.get(MaintenanceTaskTemplate, task_id)
    if not template or template.owner_user_id != current_user.id:
        raise HTTPException(status_code=404, detail='Maintenance task not found')
    return _maintenance_task_impact(template, current_user, db)


@router.delete('/maintenance-tasks/{task_id}', response_model=MaintenanceTaskDeleteImpact)
def delete_maintenance_task(task_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    template = db.get(MaintenanceTaskTemplate, task_id)
    if not template or template.owner_user_id != current_user.id:
        raise HTTPException(status_code=404, detail='Maintenance task not found')

    impact = _maintenance_task_impact(template, current_user, db)
    matching_assets = db.scalars(
        select(Asset).where(
            Asset.owner_user_id == current_user.id,
            Asset.asset_type == template.asset_type,
        )
    ).all()
    matching_asset_ids = {asset.id for asset in matching_assets}

    if matching_asset_ids:
        schedules = db.scalars(
            select(MaintenanceSchedule).where(MaintenanceSchedule.asset_id.in_(matching_asset_ids))
        ).all()
        for schedule in schedules:
            if schedule.title.strip().lower() == template.task_name.strip().lower():
                schedule.active = False

        events = db.scalars(
            select(MaintenanceEvent).where(MaintenanceEvent.asset_id.in_(matching_asset_ids))
        ).all()
        for event in events:
            tasks = [part.strip() for part in event.event_type.split(',') if part.strip()]
            kept_tasks = [task for task in tasks if task.lower() != template.task_name.strip().lower()]
            if len(kept_tasks) == len(tasks):
                continue
            if not kept_tasks:
                db.delete(event)
            else:
                event.event_type = ', '.join(kept_tasks)

    db.delete(template)
    db.commit()
    return impact


@router.put('/maintenance-tasks/rename')
def rename_maintenance_task(payload: MaintenanceTaskRename, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    old_name = payload.old_name.strip()
    new_name = payload.new_name.strip()
    if not old_name or not new_name:
        raise HTTPException(status_code=400, detail='Both old and new task names are required')

    events = db.scalars(
        select(MaintenanceEvent)
        .where(MaintenanceEvent.performed_by_user_id == current_user.id)
    ).all()
    updated_events = 0
    for event in events:
        updated = False
        tasks = [part.strip() for part in event.event_type.split(',')]
        normalized_tasks: list[str] = []
        for task in tasks:
            if not task:
                continue
            if task.lower() == old_name.lower():
                normalized_tasks.append(new_name)
                updated = True
            else:
                normalized_tasks.append(task)
        if updated:
            event.event_type = ', '.join(normalized_tasks)
            updated_events += 1
    templates = db.scalars(
        select(MaintenanceTaskTemplate).where(MaintenanceTaskTemplate.owner_user_id == current_user.id)
    ).all()
    for template in templates:
        if template.task_name.strip().lower() == old_name.lower():
            template.task_name = new_name

    db.commit()
    return {'updated_events': updated_events}


@router.get('/dashboard', response_model=DashboardOut)
def dashboard(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    assets = db.scalars(select(Asset).where(Asset.owner_user_id == current_user.id, Asset.archived_at.is_(None))).all()
    due_soon: list[DashboardItem] = []
    overdue: list[DashboardItem] = []

    for asset in assets:
        meter_map = latest_meter_map(asset.meters)
        for schedule in asset.schedules:
            if not schedule.active:
                continue
            asset_events = db.scalars(
                select(MaintenanceEvent)
                .where(MaintenanceEvent.asset_id == asset.id)
                .order_by(desc(MaintenanceEvent.performed_at))
            ).all()
            last_event = next((
                event for event in asset_events
                if schedule.title.strip().lower() in [task.strip().lower() for task in event.event_type.split(',')]
            ), None)
            if last_event is None:
                status_value = 'overdue'
            else:
                status_value = evaluate_schedule_status(
                    schedule,
                    asset,
                    meter_map,
                    last_event,
                    datetime.utcnow(),
                    due_soon_window_days=current_user.upcoming_task_window_days,
                )
            if status_value == 'due_soon' and schedule.interval_days and last_event:
                due_date = last_event.performed_at + timedelta(days=schedule.interval_days)
                if due_date > datetime.utcnow() + timedelta(days=current_user.upcoming_task_window_days):
                    status_value = 'not_due'
            item = DashboardItem(
                asset_id=asset.id,
                schedule_id=schedule.id,
                asset_name=asset.name,
                thumbnail_path=asset.thumbnail_path,
                schedule_title=schedule.title,
                status=status_value,
            )
            if status_value == 'overdue':
                overdue.append(item)
            elif status_value == 'due_soon':
                due_soon.append(item)

    recent_events = db.execute(
        select(MaintenanceEvent, Asset.name, Asset.service_trigger)
        .join(Asset, Asset.id == MaintenanceEvent.asset_id)
        .where(Asset.owner_user_id == current_user.id)
        .order_by(desc(MaintenanceEvent.performed_at))
        .limit(10)
    ).all()
    recent_items = [
        DashboardOut.RecentEventItem(
            id=event.id,
            asset_id=event.asset_id,
            asset_name=asset_name,
            service_trigger=service_trigger,
            performed_at=event.performed_at,
            event_type=event.event_type,
            notes=event.notes,
            completion_meter_value=event.completion_meter_value,
        )
        for event, asset_name, service_trigger in recent_events
    ]
    return DashboardOut(due_soon=due_soon, overdue=overdue, recent_events=recent_items)


@router.get('/alerts')
def alerts(data: DashboardOut = Depends(dashboard)):
    return {'due_soon': data.due_soon, 'overdue': data.overdue}
